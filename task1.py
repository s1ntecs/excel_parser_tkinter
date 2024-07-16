import openpyxl
from typing import List, Optional, Union
from openpyxl.utils import column_index_from_string, get_column_letter


""" Реализация класса согласно Задания:
    Создайте класс, который принимает на вход диапазон ячеек Excel, A1:D35,
    также класс должен уметь хранить полученные данные в виде массива в
    переменной My_table, также класс должен иметь методы:
        get_cell_value – возвращает значение из массива по адресу;
        get_row – возвращает целую строку из массива
        get_column – возвращает целый столбец из массива
        find_word – возвращает адреса ячеек массива которые
            содержать слово «нет данных»
"""


class ExcelRangeProcessor:
    def __init__(self, file_path: str, range_str: str) -> None:
        self.file_path: str = file_path
        self.range_str: str = range_str
        self.My_table: List[List[Union[str, None]]] = self._load_data()

    def _load_data(self) -> List[List[Union[str, None]]]:
        """Загрузить данные из диапазона ячеек."""
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        data: List[List[Union[str, None]]] = []

        for row in ws[self.range_str]:
            data_row = []
            for cell in row:
                data_row.append(cell.value)
            data.append(data_row)

        return data

    def get_cell_value(self, cell_address: str) -> Optional[Union[str, None]]:
        """Получает значение ячейки из массива по адресу."""
        col_letter = cell_address[0]
        row_number = int(cell_address[1:]) - 1

        col_index = column_index_from_string(col_letter) - 1
        try:
            return self.My_table[row_number][col_index]
        except IndexError:
            return None

    def get_row(self, row_number: int) -> List[Optional[Union[str, None]]]:
        """Возвращает целую строку из массива."""
        try:
            return self.My_table[row_number - 1]
        except IndexError:
            return []

    def get_column(self, col_letter: str) -> List[Optional[Union[str, None]]]:
        """Возвращает целый столбец из массива."""
        col_index = column_index_from_string(col_letter) - 1
        column_data = []
        for row in self.My_table:
            try:
                column_data.append(row[col_index])
            except IndexError:
                column_data.append(None)
        return column_data

    def find_word(self, word: str) -> List[str]:
        """Возвращает адреса ячеек, которые содержат заданное слово."""
        addresses = []
        for row_idx, row in enumerate(self.My_table):
            for col_idx, cell in enumerate(row):
                if cell == word:
                    cell_address = \
                        f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                    addresses.append(cell_address)
        return addresses


# Пример использования
if __name__ == "__main__":
    processor = ExcelRangeProcessor("data/coordinates.xlsx", "A1:E35")

    cell_value = processor.get_cell_value("B2")
    row_values = processor.get_row(2)
    column_values = processor.get_column("B")
    word_addresses = processor.find_word("нет данных")

    print(f"Значение в ячейке B2: {cell_value}")
    print(f"Все значения во 2 ряде: {row_values}")
    print(f"Все значения столбца B: {column_values}")
    print(f"Адресса которые содержат 'нет данных': {word_addresses}")
