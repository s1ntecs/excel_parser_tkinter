import openpyxl
import os
from typing import List, Optional, Union

import openpyxl.utils


class ExcelRangeProcessor:
    def __init__(self, file_paths: Optional[List[str]] = None,
                 folder_path: Optional[str] = None,
                 range_str: str = "") -> None:
        self.file_paths: Optional[List[str]] = file_paths
        self.folder_path: Optional[str] = folder_path
        self.range_str: str = range_str
        self.My_table: List[Union[str, List[List[Union[str, None]]]]] = []
        if self.file_paths:
            self.load_files(file_paths)
        else:
            self.load_folder()

    def _load_data(self, file_path: str):
        """Загрузить данные из файла {file_path}.xlsx."""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data: List[List[Union[str, None]]] = []

        for row in ws[self.range_str]:
            data_row = []
            for cell in row:
                data_row.append(cell.value)
            data.append(data_row)

        return data

    def load_files(self, file_paths: List[str]) -> None:
        """Загрузить данные из всех указанных файлов .xlsx."""
        all_data: List[Union[str, List[List[Union[str, None]]]]] = []
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            file_data = self._load_data(file_path)
            # Добавляем в первую ячейку списка название excel файла
            all_data.append([file_name, file_data])
        self.My_table = all_data

    def load_folder(self) -> None:
        """Загрузить данные из всех файлов .xlsx в указанной папке."""
        if self.folder_path:
            file_paths = [
                os.path.join(self.folder_path, filename)
                for filename in os.listdir(self.folder_path)
                if filename.endswith(".xlsx")
            ]
            self.load_files(file_paths)

    def get_cell_value(self, cell_address: str) -> str:
        """Получает данные по ячейке {cell_address}"""
        col_letter = cell_address[0]
        row_number = int(cell_address[1:]) - 1

        col_index = openpyxl.utils.column_index_from_string(col_letter) - 1
        result_str = ""
        for file_table in self.My_table:
            file_name = file_table[0]
            result_str += \
                f"{file_name}: {file_table[1][row_number][col_index]}\n"
        return result_str

    def get_row(self, row_number: int) -> str:
        """Получает данные по заданной строке"""
        result_str = ""
        for file_table in self.My_table:
            file_name = file_table[0]
            result_str += \
                f"{file_name}: {file_table[1][row_number - 1]}\n"
        return result_str

    def get_column(self, col_letter: str) -> str:
        """Получает данные по заданному столбцу"""
        col_index = openpyxl.utils.column_index_from_string(col_letter) - 1
        result_str = ""
        for file_table in self.My_table:
            file_name = file_table[0]
            column_data = [row[col_index] for row in file_table[1]]
            result_str += \
                f"{file_name}: {column_data}\n"
        return result_str

    def find_word(self, word: str) -> str:
        """Ищет ячейки по соотвествующему слову"""
        result_str = ""
        for file_table in self.My_table:
            file_name = file_table[0]
            for row_idx, row in enumerate(file_table[1]):
                for col_idx, cell in enumerate(row):
                    if cell == word:
                        cell_address = \
                            f"{openpyxl.utils.get_column_letter(
                                col_idx + 1)}{row_idx + 1}"
                        result_str += \
                            f"{file_name}: {cell_address}\n"
        return result_str
